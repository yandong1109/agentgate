import asyncio
from io import BytesIO
import json
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile, ZipInfo

import openpyxl
from fastapi.testclient import TestClient
from starlette.datastructures import UploadFile

from agentgate.case import build_excel
from agentgate.domain import Case, CaseTurn, DatasetVersion
from agentgate.server.application import create_app


XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADERS = (
    "case_id", "case_name", "case_description", "category", "difficulty",
    "tags_json", "initial_state_json", "turn_id", "turn_order", "input_json",
    "expected_skill", "expectations_json", "required_tools_json", "forbidden_tools_json",
    "policy_rules_json", "turn_notes",
)


def _workbook_with_multiple_issues() -> bytes:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Cases"
    sheet.append(HEADERS)
    sheet.append((
        "case-1", "Case", "", "positive", "medium", "{", "{}", "turn-1", 1,
        None, None, "[", "[]", "[]", "[]", "",
    ))
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _zip_bytes(entries: tuple[tuple[str, bytes], ...] = ()) -> bytes:
    content = BytesIO()
    with ZipFile(content, "w", ZIP_DEFLATED) as archive:
        for name, value in entries:
            archive.writestr(name, value)
    return content.getvalue()


def _without_archive_member(content: bytes, excluded_name: str) -> bytes:
    output = BytesIO()
    with ZipFile(BytesIO(content)) as source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        for member in source.infolist():
            if member.filename != excluded_name:
                target.writestr(member, source.read(member.filename))
    return output.getvalue()


def _pad_xlsx_to_size(content: bytes, target_size: int) -> bytes:
    def with_padding(padding_size: int) -> bytes:
        output = BytesIO()
        with ZipFile(BytesIO(content)) as source, ZipFile(output, "w") as target:
            for member in source.infolist():
                target.writestr(member, source.read(member.filename))
            padding = ZipInfo("agentgate-padding.bin")
            padding.compress_type = ZIP_STORED
            target.writestr(padding, b"\0" * padding_size)
        return output.getvalue()

    overhead = len(with_padding(0))
    output = with_padding(target_size - overhead)
    assert len(output) == target_size
    return output


def _dataset_ids(client: TestClient) -> set[str]:
    return {item["id"] for item in client.get("/api/datasets").json()}


def _stored_record_counts(client: TestClient) -> tuple[int, int]:
    with client.app.state.repository._connect() as db:
        return (
            db.execute("SELECT COUNT(*) FROM datasets").fetchone()[0],
            db.execute("SELECT COUNT(*) FROM dataset_versions").fetchone()[0],
        )


def _excel_error(response) -> dict:
    detail = response.json()["detail"]
    assert set(detail) == {"code", "total_count", "truncated", "issues"}
    assert detail["total_count"] >= len(detail["issues"])
    assert detail["truncated"] == (detail["total_count"] > len(detail["issues"]))
    return detail


def _excel_issues(response) -> list[dict]:
    return _excel_error(response)["issues"]


def test_excel_import_creates_draft_and_published_export_downloads(tmp_path):
    case = Case(
        id="imported-case",
        name="Imported Case",
        turns=(CaseTurn(id="imported-turn", input={"message": "hello"}),),
    )
    workbook = build_excel(DatasetVersion(dataset_id="source", cases=(case,)))

    with TestClient(create_app(tmp_path / "dataset-excel-api.db")) as client:
        imported = client.post(
            "/api/datasets/import/excel",
            data={"name": "Imported Excel", "description": "From workbook"},
            files={
                "file": (
                    "import.xlsx",
                    workbook,
                    XLSX_MEDIA_TYPE,
                )
            },
        )

        assert imported.status_code == 201
        payload = imported.json()
        dataset_id = payload["dataset"]["id"]
        assert dataset_id != "source"
        assert payload["dataset"]["name"] == "Imported Excel"
        assert payload["version"]["status"] == "draft"
        assert payload["version"]["version"] is None
        assert payload["version"]["cases"][0]["id"] == "imported-case"
        assert payload["version"]["cases"][0]["turns"][0]["id"] == "imported-turn"

        detail = client.get(f"/api/datasets/{dataset_id}")
        assert detail.status_code == 200
        assert detail.json()["dataset"]["description"] == "From workbook"

        published = client.post(f"/api/datasets/{dataset_id}/drafts/publish")
        assert published.status_code == 200
        downloaded = client.get(
            f"/api/datasets/{dataset_id}/versions/{published.json()['version']}/export/excel"
        )

        assert downloaded.status_code == 200
        assert downloaded.headers["content-type"] == XLSX_MEDIA_TYPE
        assert downloaded.headers["content-disposition"].startswith(
            'attachment; filename="Imported-Excel-v1.xlsx"; filename*=UTF-8\'\''
        )
        assert downloaded.headers["etag"]
        assert downloaded.headers["cache-control"] == "private, immutable"
        sheet = openpyxl.load_workbook(BytesIO(downloaded.content), read_only=True)["Cases"]
        assert next(sheet.iter_rows(min_row=2, values_only=True))[0] == "imported-case"


def test_excel_template_download_contains_instructions_and_empty_cases_sheet(tmp_path):
    with TestClient(create_app(tmp_path / "excel-template-api.db")) as client:
        response = client.get("/api/datasets/excel/template")

    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_MEDIA_TYPE
    assert response.headers["content-disposition"] == (
        'attachment; filename="agentgate-dataset-template.xlsx"'
    )
    workbook = openpyxl.load_workbook(BytesIO(response.content), read_only=True)
    assert workbook.sheetnames == ["Cases", "Instructions", "Metadata"]
    assert list(workbook["Cases"].iter_rows(min_row=2, values_only=True)) == []


def test_excel_export_rejects_a_draft_without_a_published_version(tmp_path):
    workbook = build_excel(DatasetVersion(
        dataset_id="source",
        cases=(Case(name="Case", turns=(CaseTurn(input={"message": "hello"}),)),),
    ))

    with TestClient(create_app(tmp_path / "draft-excel-export-api.db")) as client:
        imported = client.post(
            "/api/datasets/import/excel",
            data={"name": "Draft dataset"},
            files={"file": ("draft.xlsx", workbook, XLSX_MEDIA_TYPE)},
        )
        assert imported.status_code == 201

        response = client.get(
            f"/api/datasets/{imported.json()['dataset']['id']}/versions/1/export/excel"
        )

        assert response.status_code == 404
        assert response.json()["detail"].startswith("unknown dataset version")


def test_excel_import_requires_a_name_and_creates_no_dataset(tmp_path):
    workbook = build_excel(DatasetVersion(
        dataset_id="source",
        cases=(Case(name="Case", turns=(CaseTurn(input={}),)),),
    ))

    with TestClient(create_app(tmp_path / "missing-name-excel-api.db")) as client:
        before = _dataset_ids(client)
        response = client.post(
            "/api/datasets/import/excel",
            files={"file": ("import.xlsx", workbook, XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 422
        assert _dataset_ids(client) == before


def test_excel_import_rejects_a_non_xlsx_filename_with_structured_issue(tmp_path):
    with TestClient(create_app(tmp_path / "wrong-extension-excel-api.db")) as client:
        before = _dataset_ids(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Wrong extension"},
            files={"file": ("import.csv", b"not a spreadsheet", "text/csv")},
        )

        assert response.status_code == 415
        assert _excel_issues(response) == [{
            "sheet": "Cases",
            "row": None,
            "column": None,
            "message": "file must have a .xlsx filename",
        }]
        assert _dataset_ids(client) == before


def test_excel_import_rejects_bodies_over_ten_mebibytes_before_creating_a_dataset(tmp_path):
    with TestClient(create_app(tmp_path / "oversized-excel-api.db")) as client:
        before = _dataset_ids(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Oversized"},
            files={
                "file": (
                    "oversized.xlsx",
                    b"x" * (10 * 1024 * 1024 + 1),
                    XLSX_MEDIA_TYPE,
                )
            },
        )

        assert response.status_code == 413
        assert _excel_issues(response)[0]["message"] == "XLSX upload exceeds 10 MiB"
        assert _dataset_ids(client) == before


def test_excel_import_accepts_a_valid_file_at_exactly_ten_mebibytes(tmp_path):
    workbook = build_excel(DatasetVersion(
        dataset_id="source",
        cases=(Case(name="Case", turns=(CaseTurn(input={"message": "hello"}),)),),
    ))
    workbook = _pad_xlsx_to_size(workbook, 10 * 1024 * 1024)

    with TestClient(create_app(tmp_path / "exact-upload-limit.db")) as client:
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Exact limit"},
            files={"file": ("exact.xlsx", workbook, XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 201
        assert response.json()["dataset"]["name"] == "Exact limit"


def test_excel_import_reads_at_most_one_byte_beyond_the_upload_limit(tmp_path, monkeypatch):
    original_read = UploadFile.read
    read_sizes: list[int] = []

    async def record_read(upload: UploadFile, size: int = -1) -> bytes:
        read_sizes.append(size)
        return await original_read(upload, size)

    monkeypatch.setattr(UploadFile, "read", record_read)

    with TestClient(create_app(tmp_path / "bounded-read-excel-api.db")) as client:
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Bounded read"},
            files={"file": ("invalid.xlsx", b"not an xlsx", XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 422
        assert read_sizes == [10 * 1024 * 1024 + 1]


def test_excel_import_rejects_oversized_content_length_before_upload_parsing(
    tmp_path, monkeypatch,
):
    read_sizes: list[int] = []

    async def record_read(_upload: UploadFile, size: int = -1) -> bytes:
        read_sizes.append(size)
        return b""

    monkeypatch.setattr(UploadFile, "read", record_read)

    with TestClient(create_app(tmp_path / "request-length-limit.db")) as client:
        response = client.post(
            "/api/datasets/import/excel",
            headers={
                "content-type": "multipart/form-data; boundary=unused",
                "content-length": str(11 * 1024 * 1024 + 1),
            },
            content=b"unused",
        )

        assert response.status_code == 413
        assert _excel_issues(response)[0]["message"] == (
            "multipart request body exceeds 11 MiB"
        )
        assert read_sizes == []


def test_excel_import_rejects_a_streamed_request_body_above_the_envelope_limit(tmp_path):
    app = create_app(tmp_path / "streamed-request-limit.db")
    sent: list[dict] = []
    chunks = iter((
        {"type": "http.request", "body": b"x" * (6 * 1024 * 1024), "more_body": True},
        {"type": "http.request", "body": b"x" * (5 * 1024 * 1024 + 1), "more_body": False},
    ))

    async def receive():
        return next(chunks)

    async def send(message):
        sent.append(message)

    scope = {
        "type": "http",
        "asgi": {"version": "3.0"},
        "http_version": "1.1",
        "method": "POST",
        "scheme": "http",
        "path": "/api/datasets/import/excel",
        "raw_path": b"/api/datasets/import/excel",
        "query_string": b"",
        "headers": [(b"content-type", b"multipart/form-data; boundary=unused")],
        "client": ("testclient", 50000),
        "server": ("testserver", 80),
    }

    asyncio.run(app(scope, receive, send))

    assert next(item for item in sent if item["type"] == "http.response.start")["status"] == 413
    response_body = b"".join(
        item.get("body", b"") for item in sent if item["type"] == "http.response.body"
    )
    detail = json.loads(response_body)["detail"]
    assert detail["issues"][0]["message"] == (
        "multipart request body exceeds 11 MiB"
    )


def test_excel_import_runs_sync_workbook_work_outside_the_async_event_loop(
    tmp_path, monkeypatch,
):
    app = create_app(tmp_path / "excel-threadpool.db")
    dataset_service = app.state.service.dataset_service
    original_import = dataset_service.import_excel
    running_loop_observed: list[bool] = []

    def record_loop_state(*args, **kwargs):
        try:
            asyncio.get_running_loop()
        except RuntimeError:
            running_loop_observed.append(False)
        else:
            running_loop_observed.append(True)
        return original_import(*args, **kwargs)

    monkeypatch.setattr(dataset_service, "import_excel", record_loop_state)
    workbook = build_excel(DatasetVersion(
        dataset_id="source",
        cases=(Case(name="Case", turns=(CaseTurn(input={"message": "hello"}),)),),
    ))

    with TestClient(app) as client:
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Threadpool import"},
            files={"file": ("threadpool.xlsx", workbook, XLSX_MEDIA_TYPE)},
        )

    assert response.status_code == 201
    assert running_loop_observed == [False]


def test_excel_import_returns_all_malformed_workbook_issues_and_creates_no_dataset(tmp_path):
    with TestClient(create_app(tmp_path / "malformed-excel-api.db")) as client:
        before = _dataset_ids(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Malformed"},
            files={
                "file": (
                    "malformed.xlsx",
                    _workbook_with_multiple_issues(),
                    XLSX_MEDIA_TYPE,
                )
            },
        )

        assert response.status_code == 422
        issues = _excel_issues(response)
        assert {issue["column"] for issue in issues} >= {
            "tags_json", "input_json", "expectations_json",
        }
        assert all(set(issue) == {"sheet", "row", "column", "message"} for issue in issues)
        assert _dataset_ids(client) == before


def test_excel_import_normalizes_an_empty_zip_package_to_structured_422(tmp_path):
    with TestClient(
        create_app(tmp_path / "empty-zip-api.db"), raise_server_exceptions=False,
    ) as client:
        before = _stored_record_counts(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Empty ZIP"},
            files={"file": ("empty.xlsx", _zip_bytes(), XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 422
        assert _excel_issues(response)[0] == {
            "sheet": "Cases",
            "row": None,
            "column": None,
            "message": "invalid XLSX content: missing [Content_Types].xml",
        }
        assert _stored_record_counts(client) == before


def test_excel_import_normalizes_a_package_missing_a_critical_part_to_422(tmp_path):
    workbook = build_excel(DatasetVersion(
        dataset_id="source",
        cases=(Case(name="Case", turns=(CaseTurn(input={"message": "hello"}),)),),
    ))
    malformed = _without_archive_member(workbook, "[Content_Types].xml")

    with TestClient(
        create_app(tmp_path / "missing-part-api.db"), raise_server_exceptions=False,
    ) as client:
        before = _stored_record_counts(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Missing part"},
            files={"file": ("missing.xlsx", malformed, XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 422
        assert _excel_issues(response)[0] == {
            "sheet": "Cases",
            "row": None,
            "column": None,
            "message": "invalid XLSX content: missing [Content_Types].xml",
        }
        assert _stored_record_counts(client) == before


def test_excel_import_api_rejects_an_archive_compression_bomb(tmp_path):
    bomb = _zip_bytes((("xl/worksheets/sheet1.xml", b"x" * 100_000),))

    with TestClient(create_app(tmp_path / "archive-bomb-api.db")) as client:
        before = _stored_record_counts(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Archive bomb"},
            files={"file": ("bomb.xlsx", bomb, XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 422
        assert "compression ratio" in _excel_issues(response)[0]["message"]
        assert _stored_record_counts(client) == before


def test_excel_import_api_returns_structured_full_validation_issues_without_saving(tmp_path):
    workbook = BytesIO()
    source = openpyxl.Workbook()
    sheet = source.active
    sheet.title = "Cases"
    sheet.append(HEADERS)
    sheet.append((
        "case-1", "Case", "", "positive", "medium", "[]", "{}", "turn-1", 1,
        "{}", None,
        '[{"id":"same","kind":"output","condition":{"kind":"equals","expected":1}},'
        '{"id":"same","kind":"output","condition":{"kind":"equals","expected":2}}]',
        '["lookup"]', '["lookup"]', "[]", "",
    ))
    source.save(workbook)

    with TestClient(create_app(tmp_path / "full-validation-excel-api.db")) as client:
        before = _stored_record_counts(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Invalid domain workbook"},
            files={"file": ("invalid.xlsx", workbook.getvalue(), XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 422
        issues = _excel_issues(response)
        assert {(item["row"], item["column"]) for item in issues} == {
            (2, "input_json"),
            (2, "required_tools_json"),
            (2, "expectations_json"),
        }
        assert all(item["sheet"] == "Cases" for item in issues)
        assert _stored_record_counts(client) == before


def test_excel_import_api_rejects_header_only_workbook_without_saving(tmp_path):
    workbook = BytesIO()
    source = openpyxl.Workbook()
    source.active.title = "Cases"
    source.active.append(HEADERS)
    source.save(workbook)

    with TestClient(create_app(tmp_path / "empty-excel-api.db")) as client:
        before = _stored_record_counts(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Empty workbook"},
            files={"file": ("empty.xlsx", workbook.getvalue(), XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 422
        assert _excel_issues(response) == [{
            "sheet": "Cases",
            "row": None,
            "column": None,
            "message": "测评集至少需要一个用例",
        }]
        assert _stored_record_counts(client) == before


def test_excel_import_api_rejects_formula_cells_with_structured_issue(tmp_path):
    workbook = BytesIO()
    source = openpyxl.Workbook()
    source.active.title = "Cases"
    source.active.append(HEADERS)
    source.active.append((
        "case-1", "Case", "", "positive", "medium", "[]", "{}", "turn-1", 1,
        "=1+1", None, "[]", "[]", "[]", "[]", "",
    ))
    source.save(workbook)

    with TestClient(create_app(tmp_path / "formula-excel-api.db")) as client:
        before = _stored_record_counts(client)
        response = client.post(
            "/api/datasets/import/excel",
            data={"name": "Formula workbook"},
            files={"file": ("formula.xlsx", workbook.getvalue(), XLSX_MEDIA_TYPE)},
        )

        assert response.status_code == 422
        assert _excel_issues(response) == [{
            "sheet": "Cases",
            "row": 2,
            "column": "input_json",
            "message": "Excel formulas are not allowed",
        }]
        assert _stored_record_counts(client) == before


def test_excel_export_api_returns_structured_422_for_unrepresentable_cells(tmp_path):
    with TestClient(create_app(tmp_path / "unrepresentable-export-api.db")) as client:
        created = client.post(
            "/api/datasets",
            json={"name": "Unrepresentable export"},
        ).json()
        dataset_id = created["dataset"]["id"]
        saved = client.post(
            f"/api/datasets/{dataset_id}/drafts/cases",
            json={
                "name": "x" * 32_768,
                "turns": [{"input": {"message": "hello"}}],
            },
        )
        assert saved.status_code == 201
        published = client.post(f"/api/datasets/{dataset_id}/drafts/publish")
        assert published.status_code == 200

        response = client.get(
            f"/api/datasets/{dataset_id}/versions/1/export/excel"
        )

        assert response.status_code == 422
        assert _excel_issues(response) == [{
            "sheet": "Cases",
            "row": 2,
            "column": "case_name",
            "message": "cell text exceeds Excel's 32,767 character limit",
        }]


def test_excel_export_returns_not_found_for_an_unknown_published_version(tmp_path):
    workbook = build_excel(DatasetVersion(
        dataset_id="source",
        cases=(Case(name="Case", turns=(CaseTurn(input={"message": "hello"}),)),),
    ))

    with TestClient(create_app(tmp_path / "unknown-version-excel-api.db")) as client:
        imported = client.post(
            "/api/datasets/import/excel",
            data={"name": "Published dataset"},
            files={"file": ("published.xlsx", workbook, XLSX_MEDIA_TYPE)},
        )
        dataset_id = imported.json()["dataset"]["id"]
        assert client.post(f"/api/datasets/{dataset_id}/drafts/publish").status_code == 200

        response = client.get(f"/api/datasets/{dataset_id}/versions/2/export/excel")

        assert response.status_code == 404
        assert response.json()["detail"].startswith("unknown dataset version")
