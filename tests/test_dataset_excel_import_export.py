from datetime import UTC, datetime
from io import BytesIO
import sqlite3
from zipfile import ZIP_DEFLATED, ZIP_STORED, ZipFile

import openpyxl
import pytest

import agentgate.case.excel_import_export as excel_codec
from agentgate.case.excel_import_export import (
    DatasetExcelValidationError,
    ExcelImportIssue,
    build_excel,
    build_excel_template,
    parse_excel,
)
from agentgate.case import DatasetService
from agentgate.domain import (
    Case,
    CaseCategory,
    CaseDifficulty,
    CaseTurn,
    DatasetVersion,
    DatasetVersionStatus,
    Equals,
    MatchesPattern,
    OutputExpectation,
    StateExpectation,
    ToolArgumentExpectation,
)
from agentgate.storage.sqlite import SQLiteRepository


HEADERS = (
    "case_id", "case_name", "case_description", "category", "difficulty",
    "tags_json", "initial_state_json", "turn_id", "turn_order", "input_json",
    "expected_skill", "expectations_json", "required_tools_json", "forbidden_tools_json",
    "policy_rules_json", "turn_notes",
)


def test_excel_runtime_uses_defused_xml_parser():
    assert openpyxl.DEFUSEDXML


def published_version_with_all_fields() -> DatasetVersion:
    case = Case(
        id="case-001",
        name="贷款审批",
        notes="多轮审批场景",
        category=CaseCategory.BOUNDARY,
        difficulty=CaseDifficulty.HARD,
        tags=("loan", "中文"),
        initial_state={"application": {"id": "A-1", "status": "new"}},
        turns=(
            CaseTurn(
                id="turn-001",
                input={"message": "申请贷款", "amount": 50000},
                expected_skill="loan_approval",
                expectations=(StateExpectation(
                    id="expect-state",
                    path="application.status",
                    condition=Equals(expected="pending"),
                    name="状态已更新",
                ),),
                required_tools=("credit_check",),
                forbidden_tools=("approve_loan",),
                policy_rules=("high_risk_requires_review",),
                notes="先收集资料",
            ),
            CaseTurn(
                id="turn-002",
                input={"application_id": "A-1", "risk": "high"},
                expected_skill="loan_approval",
                expectations=(
                    ToolArgumentExpectation(
                        id="expect-tool",
                        tool="request_human_review",
                        path="application_id",
                        occurrence="first",
                        condition=Equals(expected="A-1"),
                    ),
                    OutputExpectation(
                        id="expect-output",
                        path="message",
                        condition=MatchesPattern(pattern="审核"),
                    ),
                ),
                required_tools=("request_human_review",),
                notes="人工审核",
            ),
        ),
    )
    return DatasetVersion(
        id="version-001",
        dataset_id="dataset-001",
        dataset_name="贷款数据集",
        version=1,
        status=DatasetVersionStatus.PUBLISHED,
        cases=(case,),
        published_at=datetime(2026, 8, 19, tzinfo=UTC),
    )


def test_excel_round_trip_preserves_multiturn_case_and_ids():
    version = published_version_with_all_fields()

    content = build_excel(version)

    workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook["Cases"]
    assert workbook.sheetnames == ["Cases", "Instructions", "Metadata"]
    assert tuple(next(sheet.iter_rows(values_only=True))) == HEADERS
    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert [row[0] for row in rows] == ["case-001", "case-001"]
    assert [row[7] for row in rows] == ["turn-001", "turn-002"]
    assert [row[8] for row in rows] == [1, 2]
    assert parse_excel(content) == version.cases


def test_excel_template_explains_human_readable_multiturn_grouping():
    workbook = openpyxl.load_workbook(
        BytesIO(build_excel_template()), read_only=True, data_only=True
    )

    assert workbook.sheetnames == ["Cases", "Instructions", "Metadata"]
    assert tuple(next(workbook["Cases"].iter_rows(values_only=True))) == HEADERS
    instructions = "\n".join(
        str(value)
        for row in workbook["Instructions"].iter_rows(values_only=True)
        for value in row
        if value
    )
    assert "loan-001" in instructions
    assert "不需要填写 UUID" in instructions
    metadata = dict(workbook["Metadata"].iter_rows(values_only=True))
    assert metadata["format"] == "agentgate.dataset.xlsx"
    assert metadata["format_version"] == "1"


def test_excel_export_keeps_formula_like_text_literal_and_lossless():
    case = Case(
        id="=1+1",
        name='=HYPERLINK("https://example.test","Case")',
        notes="=1+1",
        turns=(CaseTurn(
            id='=HYPERLINK("https://example.test","Turn")',
            input={"message": "hello"},
            expected_skill="=1+1",
            notes='=HYPERLINK("https://example.test","Notes")',
        ),),
    )
    version = DatasetVersion(dataset_id="formula-text", cases=(case,))

    content = build_excel(version)

    workbook = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=False)
    row = next(workbook["Cases"].iter_rows(min_row=2))
    formula_like_columns = (0, 1, 2, 7, 10, 15)
    assert all(row[index].data_type == "s" for index in formula_like_columns)
    assert tuple(row[index].value for index in formula_like_columns) == (
        "=1+1",
        '=HYPERLINK("https://example.test","Case")',
        "=1+1",
        '=HYPERLINK("https://example.test","Turn")',
        "=1+1",
        '=HYPERLINK("https://example.test","Notes")',
    )
    assert parse_excel(content) == version.cases


def _workbook_bytes(rows=(), headers=HEADERS, sheet_name="Cases"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    content = BytesIO()
    workbook.save(content)
    return content.getvalue()


def _valid_row(**changes):
    row = dict(zip(HEADERS, (
        "case-1", "Case", "", "positive", "medium", "[]", "{}", "turn-1", 1,
        '{"message":"hello"}', None, "[]", "[]", "[]", "[]", "",
    ), strict=True))
    row.update(changes)
    return tuple(row[header] for header in HEADERS)


def _issues(content):
    with pytest.raises(DatasetExcelValidationError) as error:
        parse_excel(content)
    return error.value.issues


def test_excel_reports_missing_sheet_and_required_header_locations():
    missing_sheet = _issues(_workbook_bytes(sheet_name="Other"))
    headers = tuple(header for header in HEADERS if header != "input_json")
    malformed_header = _issues(_workbook_bytes(headers=headers))

    assert missing_sheet[0].sheet == "Cases"
    assert missing_sheet[0].row is None
    assert missing_sheet[0].column is None
    assert malformed_header[0].sheet == "Cases"
    assert malformed_header[0].row == 1
    assert malformed_header[0].column == "input_json"


def test_excel_accepts_missing_optional_columns_and_reordered_headers():
    headers = ("input_json", "case_name")
    cases = parse_excel(_workbook_bytes((('{"message":"hello"}', "Case"),), headers))

    assert len(cases) == 1
    case = cases[0]
    assert case.id
    assert case.name == "Case"
    assert case.notes == ""
    assert case.category == CaseCategory.POSITIVE
    assert case.difficulty == CaseDifficulty.MEDIUM
    assert case.tags == ()
    assert case.initial_state == {}
    assert len(case.turns) == 1
    assert case.turns[0].id
    assert case.turns[0].input == {"message": "hello"}
    assert case.turns[0].expectations == ()


def test_excel_defaults_blank_optional_cells_and_generates_ids():
    cases = parse_excel(_workbook_bytes((
        _valid_row(
            case_id=None,
            case_description=None,
            category=None,
            difficulty=None,
            tags_json=None,
            initial_state_json=None,
            turn_id=None,
            turn_order=None,
            expected_skill=None,
            expectations_json=None,
            required_tools_json=None,
            forbidden_tools_json=None,
            policy_rules_json=None,
            turn_notes=None,
        ),
    )))

    assert len(cases) == 1
    assert cases[0].id
    assert cases[0].turns[0].id
    assert cases[0].turns[0].expectations == ()


def test_excel_infers_all_blank_multiturn_orders_from_row_order():
    cases = parse_excel(_workbook_bytes((
        _valid_row(turn_id="first", turn_order=None, input_json='{"step":1}'),
        _valid_row(turn_id="second", turn_order=None, input_json='{"step":2}'),
    )))

    assert [turn.id for turn in cases[0].turns] == ["first", "second"]


def test_excel_rejects_ambiguous_repeated_anonymous_case_names():
    issues = _issues(_workbook_bytes((
        _valid_row(case_id=None, case_name="Loan", turn_id=None, turn_order=None),
        _valid_row(case_id=None, case_name="Loan", turn_id=None, turn_order=None),
    )))

    assert (3, "case_id", "repeated case_name requires a case_id for multi-turn grouping") in {
        (issue.row, issue.column, issue.message) for issue in issues
    }


def test_excel_rejects_anonymous_row_with_multiturn_order():
    issues = _issues(_workbook_bytes((
        _valid_row(case_id=None, turn_id=None, turn_order=2),
    )))

    assert (2, "case_id", "case_id is required when turn_order is greater than 1") in {
        (issue.row, issue.column, issue.message) for issue in issues
    }


def test_excel_ignores_completely_blank_rows():
    cases = parse_excel(_workbook_bytes((
        _valid_row(),
        (None,) * len(HEADERS),
    )))

    assert len(cases) == 1


def test_excel_rejects_partially_blank_multiturn_orders():
    issues = _issues(_workbook_bytes((
        _valid_row(turn_id="first", turn_order=1),
        _valid_row(turn_id="second", turn_order=None),
    )))

    assert (3, "turn_order", "value is required when another turn order is provided") in {
        (issue.row, issue.column, issue.message) for issue in issues
    }


def test_excel_aggregates_blank_and_malformed_json_cells():
    issues = _issues(_workbook_bytes((
        _valid_row(input_json=None, tags_json="{"),
        _valid_row(case_id="case-2", turn_id="turn-2", expectations_json="["),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (2, "input_json"), (2, "tags_json"), (3, "expectations_json"),
    }
    assert all(issue.sheet == "Cases" and issue.message for issue in issues)


def test_excel_rejects_formula_cells_instead_of_using_cached_results():
    issues = _issues(_workbook_bytes((
        _valid_row(
            input_json="=1+1",
            expected_skill='=HYPERLINK("https://example.test","Skill")',
        ),
    )))

    formula_issues = {
        (issue.row, issue.column, issue.message)
        for issue in issues
        if "formula" in issue.message.lower()
    }
    assert formula_issues == {
        (2, "input_json", "Excel formulas are not allowed"),
        (2, "expected_skill", "Excel formulas are not allowed"),
    }


def test_excel_reports_conflicting_case_fields_and_duplicate_turn_ids():
    issues = _issues(_workbook_bytes((
        _valid_row(),
        _valid_row(case_name="Changed", turn_order=2),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (3, "case_name"), (3, "turn_id"),
    }


def test_excel_reports_duplicate_and_noncontiguous_turn_orders():
    issues = _issues(_workbook_bytes((
        _valid_row(turn_id="turn-1", turn_order=1),
        _valid_row(turn_id="turn-2", turn_order=1),
        _valid_row(turn_id="turn-3", turn_order=3),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (3, "turn_order"), (4, "turn_order"),
    }


def test_excel_reports_invalid_case_enums_and_expectations():
    issues = _issues(_workbook_bytes((
        _valid_row(
            category="unknown",
            difficulty="impossible",
            expectations_json='[{"kind":"state"}]',
        ),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (2, "category"), (2, "difficulty"), (2, "expectations_json"),
    }


def test_excel_reports_row_limit_and_invalid_xlsx_content():
    content = _workbook_bytes([
        _valid_row(turn_id=f"turn-{index}") for index in range(10_001)
    ])
    row_limit = _issues(content)
    invalid_xlsx = _issues(b"not an xlsx")

    assert row_limit[0].sheet == "Cases"
    assert row_limit[0].row == 10_002
    assert row_limit[0].column is None
    assert invalid_xlsx[0].sheet == "Cases"
    assert invalid_xlsx[0].row is None
    assert invalid_xlsx[0].column is None


def test_excel_reports_input_byte_limit_before_reading_workbook():
    issues = _issues(b"x" * (10 * 1024 * 1024 + 1))

    assert issues[0].sheet == "Cases"
    assert issues[0].row is None
    assert issues[0].column is None


def test_excel_export_preserves_a_cell_at_the_excel_character_limit():
    name = "x" * 32_767
    version = DatasetVersion(
        dataset_id="cell-limit",
        cases=(Case(name=name, turns=(CaseTurn(input={"message": "hello"}),)),),
    )

    assert parse_excel(build_excel(version))[0].name == name


def test_excel_export_rejects_a_cell_above_the_excel_character_limit():
    version = DatasetVersion(
        dataset_id="cell-too-long",
        cases=(Case(
            name="x" * 32_768,
            turns=(CaseTurn(input={"message": "hello"}),),
        ),),
    )

    with pytest.raises(ValueError) as error:
        build_excel(version)

    assert type(error.value).__name__ == "DatasetExcelExportError"
    issue = error.value.issues[0]
    assert (issue.sheet, issue.row, issue.column) == ("Cases", 2, "case_name")
    assert "32,767" in issue.message


def test_excel_export_rejects_xml_illegal_control_characters_as_a_structured_error():
    version = DatasetVersion(
        dataset_id="illegal-xml",
        cases=(Case(
            name="bad\x01name",
            turns=(CaseTurn(input={"message": "hello"}),),
        ),),
    )

    with pytest.raises(ValueError) as error:
        build_excel(version)

    assert type(error.value).__name__ == "DatasetExcelExportError"
    issue = error.value.issues[0]
    assert (issue.sheet, issue.row, issue.column) == ("Cases", 2, "case_name")
    assert "XML" in issue.message


def _corrupt_cases_xml(content):
    corrupted = BytesIO()
    with ZipFile(BytesIO(content)) as source, ZipFile(corrupted, "w", ZIP_DEFLATED) as target:
        for member in source.infolist():
            data = source.read(member.filename)
            if member.filename == "xl/worksheets/sheet1.xml":
                data = b"<worksheet"
            target.writestr(member, data)
    return corrupted.getvalue()


def _zip_bytes(entries: tuple[tuple[str, bytes], ...], compression=ZIP_DEFLATED) -> bytes:
    content = BytesIO()
    with ZipFile(content, "w", compression) as archive:
        for name, value in entries:
            archive.writestr(name, value)
    return content.getvalue()


def test_excel_rejects_an_archive_with_an_oversized_uncompressed_entry(monkeypatch):
    monkeypatch.setattr(
        excel_codec, "MAX_XLSX_ENTRY_UNCOMPRESSED_BYTES", 100,
    )
    content = _zip_bytes((("xl/worksheets/sheet1.xml", b"x" * 101),), ZIP_STORED)

    issues = _issues(content)

    assert any("single-entry uncompressed limit" in issue.message for issue in issues)


def test_excel_rejects_an_archive_above_the_total_uncompressed_limit(monkeypatch):
    monkeypatch.setattr(
        excel_codec, "MAX_XLSX_UNCOMPRESSED_BYTES", 150,
    )
    content = _zip_bytes((
        ("xl/worksheets/sheet1.xml", b"x" * 80),
        ("xl/sharedStrings.xml", b"y" * 80),
    ), ZIP_STORED)

    issues = _issues(content)

    assert any("total uncompressed limit" in issue.message for issue in issues)


def test_excel_rejects_an_archive_entry_with_an_extreme_compression_ratio():
    content = _zip_bytes((("xl/worksheets/sheet1.xml", b"x" * 100_000),))

    issues = _issues(content)

    assert any("compression ratio" in issue.message for issue in issues)


def test_excel_rejects_an_archive_with_too_many_entries(monkeypatch):
    monkeypatch.setattr(excel_codec, "MAX_XLSX_ENTRIES", 2)
    content = _zip_bytes((
        ("[Content_Types].xml", b"types"),
        ("xl/workbook.xml", b"workbook"),
        ("xl/worksheets/sheet1.xml", b"sheet"),
    ), ZIP_STORED)

    issues = _issues(content)

    assert any("archive entry count" in issue.message for issue in issues)


def test_excel_rejects_external_links_and_embedded_active_content():
    content = BytesIO()
    with ZipFile(BytesIO(_workbook_bytes()), "r") as source, ZipFile(
        content, "w", ZIP_DEFLATED
    ) as target:
        for member in source.infolist():
            target.writestr(member, source.read(member.filename))
        target.writestr("xl/externalLinks/externalLink1.xml", b"<externalLink/>")
        target.writestr("xl/embeddings/oleObject1.bin", b"embedded")

    issues = _issues(content.getvalue())

    assert any("external links" in issue.message for issue in issues)
    assert any("embedded objects" in issue.message for issue in issues)


def test_excel_bounds_retained_import_issues(monkeypatch):
    monkeypatch.setattr(excel_codec, "MAX_IMPORT_ISSUES", 2)
    content = _workbook_bytes((
        _valid_row(case_id="case-1", input_json="{"),
        _valid_row(case_id="case-2", turn_id="turn-2", input_json="{"),
        _valid_row(case_id="case-3", turn_id="turn-3", input_json="{"),
    ))

    with pytest.raises(DatasetExcelValidationError) as error:
        parse_excel(content)

    assert len(error.value.issues) == 2
    assert error.value.total_count >= 3
    assert error.value.truncated


def test_excel_wraps_lazy_worksheet_xml_errors_as_import_issues():
    issues = _issues(_corrupt_cases_xml(_workbook_bytes()))

    assert issues[0].sheet == "Cases"
    assert issues[0].row is None
    assert issues[0].column is None
    assert issues[0].message.startswith("invalid XLSX content")


def test_excel_does_not_normalize_application_value_errors_as_package_errors(monkeypatch):
    def raise_application_error(*_args, **_kwargs):
        raise ValueError("application parsing defect")

    monkeypatch.setattr(excel_codec, "_build_case", raise_application_error)

    with pytest.raises(ValueError, match="application parsing defect") as error:
        parse_excel(_workbook_bytes((_valid_row(),)))

    assert not isinstance(error.value, DatasetExcelValidationError)


def test_excel_aggregates_model_errors_with_generated_case_id():
    issues = _issues(_workbook_bytes((
        _valid_row(
            case_id=None,
            category="unknown",
            difficulty="impossible",
            tags_json="{}",
            initial_state_json="[]",
            input_json="[]",
        ),
    )))

    assert {(issue.row, issue.column) for issue in issues} >= {
        (2, "category"),
        (2, "difficulty"),
        (2, "tags_json"),
        (2, "initial_state_json"),
        (2, "input_json"),
    }


def test_service_import_excel_creates_trimmed_dataset_with_unpublished_draft(tmp_path):
    service = DatasetService(SQLiteRepository(tmp_path / "import.db"))
    cases = published_version_with_all_fields().cases

    dataset, draft = service.import_excel(
        build_excel(published_version_with_all_fields()),
        "  Imported dataset  ",
        "  Imported description  ",
    )

    assert dataset.name == "Imported dataset"
    assert dataset.description == "Imported description"
    assert draft.dataset_id == dataset.id
    assert draft.status == DatasetVersionStatus.DRAFT
    assert draft.version is None
    assert draft.cases == cases
    assert service.get_draft(dataset.id) == draft
    with pytest.raises(ValueError, match="unknown dataset version"):
        service.get_version(dataset.id, 1)


def test_service_import_excel_rejects_blank_dataset_name(tmp_path):
    service = DatasetService(SQLiteRepository(tmp_path / "blank-name.db"))

    with pytest.raises(ValueError, match="dataset name is required"):
        service.import_excel(build_excel(published_version_with_all_fields()), "  ")

    assert service.list_datasets(include_archived=True) == []


def test_service_import_excel_malformed_workbook_creates_no_records(tmp_path):
    repository = SQLiteRepository(tmp_path / "malformed.db")
    service = DatasetService(repository)

    with pytest.raises(DatasetExcelValidationError):
        service.import_excel(b"not an xlsx", "Imported dataset")

    assert service.list_datasets(include_archived=True) == []
    with repository._connect() as db:
        assert db.execute("SELECT COUNT(*) FROM dataset_versions").fetchone()[0] == 0


def _repository_record_counts(repository: SQLiteRepository) -> tuple[int, int]:
    with repository._connect() as db:
        return (
            db.execute("SELECT COUNT(*) FROM datasets").fetchone()[0],
            db.execute("SELECT COUNT(*) FROM dataset_versions").fetchone()[0],
        )


def test_service_import_excel_rejects_an_empty_workbook_before_persistence(tmp_path):
    repository = SQLiteRepository(tmp_path / "empty-workbook.db")
    service = DatasetService(repository)

    with pytest.raises(DatasetExcelValidationError) as error:
        service.import_excel(_workbook_bytes(), "Empty workbook")

    assert error.value.issues == (
        ExcelImportIssue(
            sheet="Cases",
            row=None,
            column=None,
            message="测评集至少需要一个用例",
        ),
    )
    assert _repository_record_counts(repository) == (0, 0)


def test_service_import_excel_maps_empty_input_validation_to_its_source_cell(tmp_path):
    repository = SQLiteRepository(tmp_path / "empty-input.db")
    service = DatasetService(repository)

    with pytest.raises(DatasetExcelValidationError) as error:
        service.import_excel(
            _workbook_bytes((_valid_row(input_json="{}"),)),
            "Empty input",
        )

    issue = error.value.issues[0]
    assert (issue.sheet, issue.row, issue.column) == ("Cases", 2, "input_json")
    assert issue.message == "每一轮必须包含输入"
    assert _repository_record_counts(repository) == (0, 0)


def test_service_import_excel_maps_tool_overlap_to_required_tools_cell(tmp_path):
    repository = SQLiteRepository(tmp_path / "tool-overlap.db")
    service = DatasetService(repository)

    with pytest.raises(DatasetExcelValidationError) as error:
        service.import_excel(
            _workbook_bytes((_valid_row(
                required_tools_json='["lookup"]',
                forbidden_tools_json='["lookup"]',
            ),)),
            "Tool overlap",
        )

    issue = error.value.issues[0]
    assert (issue.sheet, issue.row, issue.column) == (
        "Cases", 2, "required_tools_json",
    )
    assert "lookup" in issue.message
    assert _repository_record_counts(repository) == (0, 0)


def test_service_import_excel_maps_duplicate_expectation_ids_to_expectations_cell(tmp_path):
    repository = SQLiteRepository(tmp_path / "duplicate-expectation.db")
    service = DatasetService(repository)
    first_expectation = (
        '[{"id":"duplicate","kind":"output","condition":'
        '{"kind":"equals","expected":"one"}}]'
    )
    second_expectation = (
        '[{"id":"duplicate","kind":"output","condition":'
        '{"kind":"equals","expected":"two"}}]'
    )

    with pytest.raises(DatasetExcelValidationError) as error:
        service.import_excel(
            _workbook_bytes((
                _valid_row(expectations_json=first_expectation),
                _valid_row(
                    turn_id="turn-2",
                    turn_order=2,
                    expectations_json=second_expectation,
                ),
            )),
            "Duplicate expectations",
        )

    issue = error.value.issues[0]
    assert (issue.sheet, issue.row, issue.column) == (
        "Cases", 3, "expectations_json",
    )
    assert issue.message == "同一用例内期望 ID 必须唯一"
    assert _repository_record_counts(repository) == (0, 0)


def test_service_import_excel_uses_original_rows_after_case_and_turn_reordering(tmp_path):
    repository = SQLiteRepository(tmp_path / "source-rows.db")
    service = DatasetService(repository)

    with pytest.raises(DatasetExcelValidationError) as error:
        service.import_excel(
            _workbook_bytes((
                _valid_row(
                    case_id="case-a", case_name="A", turn_id="a-2", turn_order=2,
                ),
                _valid_row(
                    case_id="case-b", case_name="B", turn_id="b-1", turn_order=1,
                    input_json="{}",
                ),
                _valid_row(
                    case_id="case-a", case_name="A", turn_id="a-1", turn_order=1,
                    input_json="{}",
                ),
            )),
            "Source rows",
        )

    empty_input_rows = {
        issue.row for issue in error.value.issues if issue.message == "每一轮必须包含输入"
    }
    assert empty_input_rows == {3, 4}
    assert _repository_record_counts(repository) == (0, 0)


def test_service_import_excel_preserves_supported_json_schema_expectation(tmp_path):
    repository = SQLiteRepository(tmp_path / "json-schema.db")
    service = DatasetService(repository)
    expectations = (
        '[{"id":"schema","kind":"output","condition":'
        '{"kind":"matches_json_schema","json_schema":{"type":"string"}}}]'
    )

    _, draft = service.import_excel(
        _workbook_bytes((_valid_row(expectations_json=expectations),)),
        "Supported schema",
    )

    condition = draft.cases[0].turns[0].expectations[0].condition
    assert condition.kind == "matches_json_schema"
    assert condition.json_schema == {"type": "string"}
    assert _repository_record_counts(repository) == (1, 1)


def test_service_import_excel_rolls_back_dataset_when_draft_save_fails(tmp_path):
    repository = SQLiteRepository(tmp_path / "rollback.db")
    service = DatasetService(repository)
    with repository._connect() as db:
        db.execute(
            """
            CREATE TRIGGER fail_draft_insert
            BEFORE INSERT ON dataset_versions
            BEGIN
                SELECT RAISE(FAIL, 'draft save failed');
            END
            """
        )

    with pytest.raises(sqlite3.IntegrityError, match="draft save failed"):
        service.import_excel(
            build_excel(published_version_with_all_fields()), "Imported dataset"
        )

    assert service.list_datasets(include_archived=True) == []
    with repository._connect() as db:
        assert db.execute("SELECT COUNT(*) FROM dataset_versions").fetchone()[0] == 0


def test_service_export_excel_rejects_draft(tmp_path):
    service = DatasetService(SQLiteRepository(tmp_path / "draft-export.db"))
    dataset, _ = service.import_excel(
        build_excel(published_version_with_all_fields()), "Imported dataset"
    )

    with pytest.raises(ValueError):
        service.export_excel(dataset.id, 1)


def test_service_export_excel_round_trips_published_version(tmp_path):
    service = DatasetService(SQLiteRepository(tmp_path / "published-export.db"))
    cases = published_version_with_all_fields().cases
    dataset, _ = service.import_excel(
        build_excel(published_version_with_all_fields()), "Imported dataset"
    )
    published = service.publish_draft(dataset.id)

    assert parse_excel(service.export_excel(dataset.id, published.version)) == cases
